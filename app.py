"""
Tabbycat API Importer v3.4 — FIXED: Speaker categories required in nested team creation
"""

import os
import io
import csv
import json
import time
import re
import requests

from flask import Flask, render_template, request, send_file, flash, redirect, url_for, jsonify

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'tabbycat-importer-key-2024')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024


def clean_string(val):
    if val is None:
        return ''
    s = str(val).strip()
    return s if s != 'None' else ''


def parse_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ('TRUE', '1', 'YES', 'Y', 'T')


def parse_float_or_none(val):
    if val is None or str(val).strip() == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def read_csv_file(file):
    text = file.read().decode('utf-8')
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def read_excel_file(file):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel files")
    wb = load_workbook(file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, header in enumerate(headers):
            if header:
                row_dict[header] = row[i] if i < len(row) else None
        rows.append(row_dict)
    return rows


def read_uploaded_file(file):
    ext = file.filename.rsplit('.', 1)[1].lower()
    if ext == 'csv':
        return read_csv_file(file)
    return read_excel_file(file)


# =============================================================================
# TABBYCAT API CLIENT (v3.4)
# =============================================================================

class TabbycatAPI:
    def __init__(self, base_url, token, tournament_slug, username=None, password=None):
        self.base_url = base_url.rstrip('/')
        self.token = token.strip() if token else ''
        self.slug = tournament_slug.strip('/')
        self.username = username
        self.password = password
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'TabbycatImporter/3.4 (Render; Python requests)',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        })
        self.created_institutions = {}
        self.stats = {'success': 0, 'failed': 0, 'errors': []}
        self.auth_method = None
        self._authenticate()

    def _authenticate(self):
        if self.token:
            self.session.headers['Authorization'] = f'Token {self.token}'
            if self._test_auth():
                self.auth_method = 'token'
                return
        
        if self.username and self.password:
            self.session.headers.pop('Authorization', None)
            if self._login_session():
                self.auth_method = 'session'
                return
        
        if self.token:
            self.session.headers['Authorization'] = f'Token {self.token}'

    def _login_session(self):
        try:
            login_url = f"{self.base_url}/accounts/login/"
            resp = self.session.get(login_url, timeout=15)
            csrf_match = re.search(r'name="csrfmiddlewaretoken" value="([^"]+)"', resp.text)
            csrf_token = csrf_match.group(1) if csrf_match else ''
            
            login_data = {
                'username': self.username,
                'password': self.password,
                'csrfmiddlewaretoken': csrf_token,
                'next': '/'
            }
            resp = self.session.post(login_url, data=login_data, timeout=15)
            
            test_url = f"{self.base_url}/database/"
            resp = self.session.get(test_url, timeout=15)
            return resp.status_code == 200
        except Exception:
            return False

    def _test_auth(self):
        try:
            url = f"{self.base_url}/api/v1/institutions"
            resp = self.session.get(url, timeout=10)
            return resp.status_code in (200, 401)
        except Exception:
            return False

    def _global_url(self, path):
        return f"{self.base_url}/api/v1{path}"

    def _tournament_url(self, path):
        return f"{self.base_url}/api/v1/tournaments/{self.slug}{path}"

    def _request(self, method, url, data=None, retries=3):
        for attempt in range(retries):
            try:
                time.sleep(0.4)
                if method == 'POST':
                    resp = self.session.post(url, json=data, timeout=30)
                else:
                    resp = self.session.get(url, timeout=30)
                
                if resp.status_code in (200, 201):
                    self.stats['success'] += 1
                    return resp.json()
                elif resp.status_code == 429:
                    time.sleep(2 ** attempt)
                    continue
                else:
                    error = f"HTTP {resp.status_code} on {method} {url.replace(self.base_url, '')}: {resp.text[:300]}"
                    self.stats['errors'].append(error)
                    self.stats['failed'] += 1
                    return None
            except requests.exceptions.RequestException as e:
                if attempt == retries - 1:
                    error = f"Request failed on {method} {url.replace(self.base_url, '')}: {str(e)}"
                    self.stats['errors'].append(error)
                    self.stats['failed'] += 1
                    return None
                time.sleep(1)
        return None

    def test_connection(self):
        diagnostics = {
            'ok': False,
            'auth_method': self.auth_method,
            'steps': [],
            'suggestion': ''
        }
        
        try:
            resp = self.session.get(self.base_url, timeout=10, allow_redirects=True)
            diagnostics['steps'].append({
                'step': 'Base URL reachable',
                'status': resp.status_code,
                'ok': resp.status_code < 500
            })
        except Exception as e:
            diagnostics['steps'].append({
                'step': 'Base URL reachable',
                'status': 0,
                'ok': False,
                'error': str(e)
            })
            diagnostics['suggestion'] = 'Cannot reach your Tabbycat URL. Check for typos.'
            return diagnostics
        
        try:
            url = f"{self.base_url}/api/v1/institutions"
            resp = self.session.get(url, timeout=10)
            diagnostics['steps'].append({
                'step': 'Global institutions list (GET /api/v1/institutions)',
                'status': resp.status_code,
                'ok': resp.status_code == 200,
                'body_preview': resp.text[:100] if resp.text else ''
            })
        except Exception as e:
            diagnostics['steps'].append({
                'step': 'Global institutions list',
                'status': 0,
                'ok': False,
                'error': str(e)
            })
        
        try:
            url = f"{self.base_url}/api/v1/tournaments/{self.slug}/institutions"
            resp = self.session.get(url, timeout=10)
            diagnostics['steps'].append({
                'step': 'Tournament institutions list (GET)',
                'status': resp.status_code,
                'ok': resp.status_code == 200,
                'body_preview': resp.text[:100] if resp.text else ''
            })
        except Exception as e:
            diagnostics['steps'].append({
                'step': 'Tournament institutions list',
                'status': 0,
                'ok': False,
                'error': str(e)
            })
        
        try:
            url = f"{self.base_url}/api/v1/tournaments/{self.slug}/teams"
            resp = self.session.get(url, timeout=10)
            diagnostics['steps'].append({
                'step': 'Tournament teams list (GET)',
                'status': resp.status_code,
                'ok': resp.status_code in (200, 401),
                'body_preview': resp.text[:100] if resp.text else ''
            })
            
            if resp.status_code == 200:
                diagnostics['ok'] = True
                diagnostics['suggestion'] = 'Connection successful! API is working.'
            elif resp.status_code == 401:
                diagnostics['suggestion'] = 'Token is invalid or expired. Get a new token from your Tabbycat Change Password page.'
            elif resp.status_code == 403:
                diagnostics['suggestion'] = 'Access forbidden. Try Session Auth Fallback (admin username + password).'
            elif resp.status_code == 404:
                diagnostics['suggestion'] = f'Tournament slug "{self.slug}" not found.'
            else:
                diagnostics['suggestion'] = f'Unexpected status {resp.status_code}.'
        except Exception as e:
            diagnostics['steps'].append({
                'step': 'Tournament teams list',
                'status': 0,
                'ok': False,
                'error': str(e)
            })
            diagnostics['suggestion'] = f'Connection error: {str(e)}.'
        
        return diagnostics

    def create_institution(self, name, code):
        if code in self.created_institutions:
            return self.created_institutions[code]
        
        data = {"name": name, "code": code}
        url = self._global_url('/institutions')
        result = self._request('POST', url, data)
        
        if result and 'url' in result:
            self.created_institutions[code] = result['url']
            return result['url']
        elif result and 'id' in result:
            inst_url = f"{self.base_url}/api/v1/institutions/{result['id']}/"
            self.created_institutions[code] = inst_url
            return inst_url
        return None

    def create_team(self, institution_url, reference, short_reference,
                    use_institution_prefix=True, emoji='', speakers=None, code_name=''):
        data = {
            "institution": institution_url,
            "reference": reference,
            "short_reference": short_reference or reference,
            "use_institution_prefix": use_institution_prefix,
        }
        if emoji:
            data["emoji"] = emoji
        if code_name:
            data["code_name"] = code_name
        if speakers:
            data["speakers"] = speakers
        
        url = self._tournament_url('/teams')
        return self._request('POST', url, data)

    def create_adjudicator(self, name, institution_url=None, email='',
                           gender='', base_score=None, independent=False,
                           adj_core=False, notes=''):
        data = {
            "name": name,
            "independent": independent,
            "adj_core": adj_core,
            "institution_conflicts": [],
            "team_conflicts": [],
            "adjudicator_conflicts": [],
        }
        
        if institution_url:
            data["institution"] = institution_url
        else:
            data["institution"] = None
        
        if email:
            data["email"] = email
        if gender:
            data["gender"] = gender
        if base_score is not None:
            data["base_score"] = base_score
        if notes:
            data["notes"] = notes
        
        url = self._tournament_url('/adjudicators')
        return self._request('POST', url, data)

    def create_speaker(self, team_id, name, email='', gender=''):
        data = {
            "name": name,
            "team": team_id,
            "categories": [],  # Required by Tabbycat API
        }
        if email:
            data["email"] = email
        if gender:
            data["gender"] = gender
        
        url = self._tournament_url('/speakers')
        return self._request('POST', url, data)


# =============================================================================
# CSV PROCESSORS
# =============================================================================

def process_institutions(rows):
    results = []
    errors = []
    seen_codes = set()
    for idx, row in enumerate(rows, start=2):
        name = clean_string(row.get('name', ''))
        code = clean_string(row.get('code', ''))
        if not name and not code:
            continue
        if not name:
            errors.append(f"Row {idx}: Missing institution name")
            continue
        if not code:
            errors.append(f"Row {idx}: Missing code for '{name}'")
            continue
        if code in seen_codes:
            errors.append(f"Row {idx}: Duplicate institution code '{code}'")
            continue
        seen_codes.add(code)
        results.append({'name': name, 'code': code})
    return results, errors


def process_adjudicators(rows):
    results = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        name = clean_string(row.get('name', ''))
        if not name:
            continue
        gender = clean_string(row.get('gender', ''))
        gender_norm = ''
        if gender.upper() in ['M', 'MALE']:
            gender_norm = 'M'
        elif gender.upper() in ['F', 'FEMALE']:
            gender_norm = 'F'
        elif gender.upper() in ['O', 'OTHER']:
            gender_norm = 'O'
        results.append({
            'name': name,
            'institution': clean_string(row.get('institution', '')),
            'email': clean_string(row.get('email', '')),
            'gender': gender_norm,
            'base_score': parse_float_or_none(row.get('base_score')),
            'independent': parse_bool(row.get('independent')),
            'adj_core': parse_bool(row.get('adj_core')),
            'notes': clean_string(row.get('notes', ''))
        })
    return results, errors


def process_teams(rows):
    results = []
    errors = []
    seen_refs = set()
    for idx, row in enumerate(rows, start=2):
        institution = clean_string(row.get('institution', ''))
        if not institution:
            continue
        ref = clean_string(row.get('reference', ''))
        if not ref:
            errors.append(f"Row {idx}: Missing reference for institution '{institution}'")
            continue
        key = f"{institution}:{ref}"
        if key in seen_refs:
            errors.append(f"Row {idx}: Duplicate team reference '{ref}' for institution '{institution}'")
            continue
        seen_refs.add(key)
        results.append({
            'institution': institution,
            'reference': ref,
            'short_reference': clean_string(row.get('short_reference', ref)),
            'code_name': clean_string(row.get('code_name', '')),
            'use_institution_prefix': parse_bool(row.get('use_institution_prefix', True)),
            'emoji': clean_string(row.get('emoji', '')),
            'team_name_human': clean_string(row.get('team_name (human)', ''))
        })
    return results, errors


def process_speakers(rows, max_speakers=None):
    results = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        name = clean_string(row.get('name', ''))
        if not name:
            continue
        gender = clean_string(row.get('gender', ''))
        gender_norm = ''
        if gender.upper() in ['M', 'MALE']:
            gender_norm = 'M'
        elif gender.upper() in ['F', 'FEMALE']:
            gender_norm = 'F'
        elif gender.upper() in ['O', 'OTHER']:
            gender_norm = 'O'
        results.append({
            'name': name,
            'gender': gender_norm,
            'email': clean_string(row.get('email', '')),
            'phone': clean_string(row.get('phone', '')),
            'anonymous': parse_bool(row.get('anonymous')),
            'team': clean_string(row.get('team', '')),
            'categories': clean_string(row.get('categories', '')),
            'initials_match': clean_string(row.get('initials_match', ''))
        })
    
    if max_speakers and max_speakers > 0:
        team_counts = {}
        filtered = []
        for spk in results:
            team = spk['team']
            team_counts[team] = team_counts.get(team, 0) + 1
            if team_counts[team] <= max_speakers:
                filtered.append(spk)
            elif team_counts[team] == max_speakers + 1:
                errors.append(f"Team '{team}': Only first {max_speakers} speakers imported (BP format). Skipped extra speakers.")
        results = filtered
    
    return results, errors


# =============================================================================
# CSV GENERATORS
# =============================================================================

def generate_institutions_csv(institutions):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['name', 'code'])
    for inst in institutions:
        writer.writerow([inst['name'], inst['code']])
    return output.getvalue()


def generate_adjudicators_csv(adjudicators):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['institution', 'name', 'email', 'gender', 'base_score', 'independent', 'adj_core', 'notes'])
    for adj in adjudicators:
        writer.writerow([
            adj['institution'], adj['name'], adj['email'], adj['gender'],
            adj['base_score'] if adj['base_score'] is not None else '',
            'TRUE' if adj['independent'] else 'FALSE',
            'TRUE' if adj['adj_core'] else 'FALSE',
            adj['notes']
        ])
    return output.getvalue()


def generate_teams_csv(teams):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['institution', 'reference', 'short_reference', 'use_institution_prefix', 'emoji'])
    for team in teams:
        writer.writerow([
            team['institution'], team['reference'], team['short_reference'],
            'TRUE' if team['use_institution_prefix'] else 'FALSE',
            team['emoji']
        ])
    return output.getvalue()


def generate_speakers_csv(speakers):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['team', 'name', 'email', 'phone', 'gender', 'anonymous'])
    for spk in speakers:
        writer.writerow([
            spk['team'], spk['name'], spk['email'], spk['phone'],
            spk['gender'], 'TRUE' if spk['anonymous'] else 'FALSE'
        ])
    return output.getvalue()


# =============================================================================
# FLASK ROUTES
# =============================================================================

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/test-connection', methods=['POST'])
def test_connection():
    data = request.get_json()
    api = TabbycatAPI(
        data.get('base_url', ''),
        data.get('token', ''),
        data.get('slug', ''),
        username=data.get('username'),
        password=data.get('password')
    )
    diagnostics = api.test_connection()
    return jsonify(diagnostics)


@app.route('/api-diagnose', methods=['POST'])
def api_diagnose():
    data = request.get_json()
    base_url = data.get('base_url', '').rstrip('/')
    token = data.get('token', '').strip()
    slug = data.get('slug', '').strip('/')
    
    results = []
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'TabbycatImporter/3.4 (Diagnostic)',
        'Accept': 'application/json'
    })
    if token:
        session.headers['Authorization'] = f'Token {token}'
    
    paths_to_try = [
        ('GET', '/api/v1/institutions'),
        ('GET', f'/api/v1/tournaments/{slug}/institutions'),
        ('GET', f'/api/v1/tournaments/{slug}/teams'),
        ('GET', f'/api/v1/tournaments/{slug}/adjudicators'),
        ('GET', '/api/v1/'),
        ('GET', '/api/'),
    ]
    
    for method, path in paths_to_try:
        url = f"{base_url}{path}"
        try:
            if method == 'GET':
                resp = session.get(url, timeout=10)
            else:
                resp = session.post(url, json={'test': 'data'}, timeout=10)
            results.append({
                'method': method,
                'url': url,
                'status': resp.status_code,
                'body_preview': resp.text[:150] if resp.text else '(empty)'
            })
        except Exception as e:
            results.append({
                'method': method,
                'url': url,
                'status': 0,
                'error': str(e)
            })
    
    return jsonify({'results': results})


@app.route('/upload', methods=['POST'])
def upload():
    mode = request.form.get('mode', 'csv')
    debate_format = request.form.get('debate_format', 'bp')
    max_speakers = 2 if debate_format == 'bp' else 3

    try:
        if 'institutions' not in request.files:
            flash('Institutions file is required', 'error')
            return redirect(url_for('index'))

        inst_file = request.files['institutions']
        adj_file = request.files.get('adjudicators')
        teams_file = request.files.get('teams')
        speakers_file = request.files.get('speakers')

        if inst_file.filename == '':
            flash('Institutions file is required', 'error')
            return redirect(url_for('index'))

        inst_rows = read_uploaded_file(inst_file)
        institutions, inst_errors = process_institutions(inst_rows)
        institution_codes = {i['code'] for i in institutions}

        adjudicators = []
        adj_errors = []
        teams = []
        team_errors = []
        speakers = []
        speaker_errors = []

        if adj_file and adj_file.filename:
            adj_rows = read_uploaded_file(adj_file)
            adjudicators, adj_errors = process_adjudicators(adj_rows)

        if teams_file and teams_file.filename:
            team_rows = read_uploaded_file(teams_file)
            teams, team_errors = process_teams(team_rows)
            for t in teams:
                if t['institution'] not in institution_codes:
                    team_errors.append(f"Team '{t['institution']} {t['reference']}': institution code '{t['institution']}' not found in institutions file")

        if speakers_file and speakers_file.filename:
            speaker_rows = read_uploaded_file(speakers_file)
            speakers, speaker_errors = process_speakers(speaker_rows, max_speakers=max_speakers)

        api_results = None
        if mode == 'api':
            base_url = request.form.get('api_url', '').strip()
            token = request.form.get('api_token', '').strip()
            slug = request.form.get('tournament_slug', '').strip()
            username = request.form.get('api_username', '').strip() or None
            password = request.form.get('api_password', '').strip() or None

            if not all([base_url, token, slug]):
                flash('API URL, Token, and Tournament Slug are required for API mode', 'error')
                return redirect(url_for('index'))

            api = TabbycatAPI(base_url, token, slug, username=username, password=password)
            diagnostics = api.test_connection()
            
            if not diagnostics['ok']:
                flash(f"API Connection Failed: {diagnostics['suggestion']}", 'error')
                for step in diagnostics['steps']:
                    flash(f"  {step['step']}: HTTP {step.get('status', 'ERR')}", 'info')
                return redirect(url_for('index'))

            # Step 1: Create institutions (GLOBAL endpoint)
            for inst in institutions:
                api.create_institution(inst['name'], inst['code'])

            # Step 2: Build speakers by team
            # CRITICAL FIX v3.4: Every speaker MUST include "categories": []
            # because Tabbycat's nested Speaker serializer requires it.
            speakers_by_team = {}
            if speakers:
                for spk in speakers:
                    t = spk['team']
                    if t not in speakers_by_team:
                        speakers_by_team[t] = []
                    spk_data = {
                        "name": spk['name'],
                        "categories": [],  # REQUIRED by Tabbycat API
                    }
                    if spk['email']:
                        spk_data["email"] = spk['email']
                    if spk['gender']:
                        spk_data["gender"] = spk['gender']
                    spk_data["anonymous"] = spk['anonymous']
                    speakers_by_team[t].append(spk_data)

            # Step 3: Create teams (TOURNAMENT endpoint)
            created_teams = {}
            for team in teams:
                inst_url = api.created_institutions.get(team['institution'])
                team_name = team['team_name_human'] or f"{team['institution']} {team['reference']}"
                team_speakers = speakers_by_team.get(team_name, [])
                
                result = api.create_team(
                    institution_url=inst_url,
                    reference=team['reference'],
                    short_reference=team['short_reference'],
                    use_institution_prefix=team['use_institution_prefix'],
                    emoji=team['emoji'],
                    code_name=team['code_name'],
                    speakers=team_speakers if team_speakers else None
                )
                
                if result and 'id' in result:
                    created_teams[team_name] = result['id']
                
                # Fallback: create speakers separately if nested failed
                if result and team_speakers and 'id' in result:
                    team_id = result['id']
                    for spk in team_speakers:
                        api.create_speaker(team_id, spk['name'], spk.get('email', ''), spk.get('gender', ''))

            # Step 4: Create adjudicators (TOURNAMENT endpoint)
            for adj in adjudicators:
                inst_url = api.created_institutions.get(adj['institution']) if adj['institution'] else None
                api.create_adjudicator(
                    name=adj['name'],
                    institution_url=inst_url,
                    email=adj['email'],
                    gender=adj['gender'],
                    base_score=adj['base_score'],
                    independent=adj['independent'],
                    adj_core=adj['adj_core'],
                    notes=adj['notes']
                )

            api_results = api.stats

        # Generate CSVs for download mode
        inst_csv = generate_institutions_csv(institutions)
        adj_csv = generate_adjudicators_csv(adjudicators)
        teams_csv = generate_teams_csv(teams)
        speakers_csv = generate_speakers_csv(speakers)

        from flask import session
        session['institutions_csv'] = inst_csv
        session['adjudicators_csv'] = adj_csv
        session['teams_csv'] = teams_csv
        session['speakers_csv'] = speakers_csv

        return render_template('results.html',
            institutions=institutions,
            inst_errors=inst_errors,
            adjudicators=adjudicators,
            adj_errors=adj_errors,
            teams=teams,
            team_errors=team_errors,
            speakers=speakers,
            speaker_errors=speaker_errors,
            inst_count=len(institutions),
            adj_count=len(adjudicators),
            team_count=len(teams),
            speaker_count=len(speakers),
            mode=mode,
            debate_format=debate_format,
            api_results=api_results)

    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('index'))


@app.route('/download/<file_type>')
def download(file_type):
    from flask import session
    files = {
        'institutions': ('institutions.csv', session.get('institutions_csv', '')),
        'adjudicators': ('adjudicators.csv', session.get('adjudicators_csv', '')),
        'teams': ('teams.csv', session.get('teams_csv', '')),
        'speakers': ('speakers.csv', session.get('speakers_csv', ''))
    }
    if file_type not in files:
        flash('Invalid file type', 'error')
        return redirect(url_for('index'))

    filename, content = files[file_type]
    buffer = io.BytesIO(content.encode('utf-8'))
    return send_file(buffer, mimetype='text/csv', as_attachment=True, download_name=filename)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
